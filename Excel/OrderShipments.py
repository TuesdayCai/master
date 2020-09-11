# By WDL 2020-4-7
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import openpyxl

'''
# 处理Date Shipped问题
'''

wb= load_workbook("ToExcel_OrderShipments.xlsx")
ws=wb["ToExcel_OrderShipments"]#也可以使用wb.get_sheet_by_name("Sheet1") 获取工作表
# 在第一列插入数据
if ws.cell(1, 1).value != 'OrderSeq':
    ws.insert_cols(0)
    for row in range(ws.max_row):
        if row == 0:
            ws.cell(row+1, 1).value = 'OrderSeq'
        else:
            ws.cell(row+1, 1).value = ws.cell(row+1, 2).value+ str(ws.cell(row+1, 5).value) + str(ws.cell(row+1, 6).value)

# 读取数据，把excel中的一个table按行读取出来，存入一个二维的list
total_list=[]
OrderSeqdict={}
for row in ws.rows:
    row_list=[]
    for cell in row:
        row_list.append(cell.value)
    total_list.append(row_list)

# 找出相同的数据 并将最大的ship date存入容器中
for i in range(len(total_list)):
    if i == 0:
        continue
    else:
        if (i+1) < len(total_list):
            if total_list[i][0] == total_list[i+1][0]:
                # 1.字符串转日期
                # print(total_list[i][18])
                a = datetime.strptime(total_list[i][17], '%m/%d/%Y %H:%M:%S %p')
                b = datetime.strptime(total_list[i+1][17], '%m/%d/%Y %H:%M:%S %p')
                diff = b-a
                if diff.days > 0:
                    maxdate = total_list[i+1][17]
                else:
                    maxdate = total_list[i][17]
                OrderSeqdict[total_list[i][0]] = maxdate

#替换ship date值 并进行颜色填充
for row in range(ws.max_row):
    if row == 0:
        continue
    else:
        for key in OrderSeqdict:
            if ws.cell(row+1, 1).value == key:
                ws.cell(row + 1, 18).value = OrderSeqdict[key]
                ws.cell(row+1,18).fill = openpyxl.styles.fills.GradientFill(stop=['FF0000', '0000FF'])
wb.save("ToExcel_OrderShipments.xlsx")
